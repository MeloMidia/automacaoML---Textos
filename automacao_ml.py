#!/usr/bin/env python3
"""
============================================================
  AUTOMAÇÃO MERCADO LIVRE — Geração de Anúncios via OpenAI
============================================================
Lê produtos da planilha no Google Drive, gera título e
descrição via API da OpenAI e salva como Google Docs na
pasta TEXTOS do cliente.

SETUP RÁPIDO:
  1. pip install -r requirements.txt
  2. Baixe credentials.json do Google Cloud Console
     (veja INSTRUCOES.txt para o passo a passo)
  3. python automacao_ml.py
============================================================
"""

import base64
import datetime
import io
import os
import pickle
import re
import time
import unicodedata
from collections import defaultdict
from pathlib import Path

# Força o diretório de trabalho para a pasta onde este script está salvo
os.chdir(Path(__file__).resolve().parent)

import openpyxl
from openai import OpenAI, APIConnectionError, APITimeoutError
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# ─────────────────────────────────────────────────────────
#  CONFIGURAÇÕES — edite aqui se necessário
# ─────────────────────────────────────────────────────────

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()

# Modelo OpenAI a usar — pode sobrescrever via OPENAI_MODEL=... no .env
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4.1-mini").strip()

# ID da pasta MERCADO LIVRE no Google Drive (fixo — não precisa alterar)
PASTA_RAIZ_ID = "1H7r7kvGIuuqZByHaAVXxvkHA64852_if"

# Linha de início dos dados na planilha (pula cabeçalhos)
LINHA_INICIO = 3  # linha 3 = primeira linha de produto

# Palavras-chave para detecção automática de colunas (sem acentos — comparação normalizada)
_HEADER_KEYWORDS = {
    "produto":   ["produto", "product", "item", "descricao", "nome"],
    "marca":     ["marca", "brand", "fabricante"],
    "codigo":    ["codigo", "code", "ref", "referencia", "sku", "cod"],
    "aplicacao": ["aplicacao", "aplicavel", "compatib", "veiculo", "aplicac"],
}


def _norm(s):
    """Remove acentos e normaliza para comparação de cabeçalhos."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', str(s).strip().lower())
        if unicodedata.category(c) != 'Mn'
    )
# Fallback caso nenhum cabeçalho seja encontrado
_COL_DEFAULTS = {"produto": 0, "marca": 1, "codigo": 2, "aplicacao": 3}

# Intervalo entre chamadas à API (segundos) para evitar rate limit
DELAY_ENTRE_PRODUTOS = 20

# ─────────────────────────────────────────────────────────

SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/documents",
]


# ══════════════════════════════════════════════════════════
#  AUTENTICAÇÃO
# ══════════════════════════════════════════════════════════

def get_credentials():
    """Autentica com Google OAuth2.

    Em servidor (deploy): lê o token de TOKEN_PICKLE_B64 (base64 do token.pickle).
    Localmente: comportamento original com arquivo token.pickle.
    """
    creds = None
    credentials_path = Path("credentials.json")
    token_path = Path("token.pickle")

    token_b64 = os.getenv("TOKEN_PICKLE_B64", "")
    if token_b64:
        creds = pickle.loads(base64.b64decode(token_b64))
    elif token_path.exists():
        with open(token_path, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            if not token_b64:
                with open(token_path, "wb") as f:
                    pickle.dump(creds, f)
        else:
            if not credentials_path.exists():
                raise SystemExit(
                    "credentials.json não encontrado e TOKEN_PICKLE_B64 não definido."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), SCOPES)
            creds = flow.run_local_server(port=0)
            with open(token_path, "wb") as f:
                pickle.dump(creds, f)

    return creds


# ══════════════════════════════════════════════════════════
#  FUNÇÕES DO GOOGLE DRIVE
# ══════════════════════════════════════════════════════════

def find_folder_by_name(drive, name, parent_id):
    """Busca uma pasta pelo nome dentro de uma pasta pai."""
    q = f"mimeType='application/vnd.google-apps.folder' and name='{name}' and '{parent_id}' in parents and trashed=false"
    r = drive.files().list(q=q, fields="files(id,name)").execute()
    files = r.get("files", [])
    return files[0] if files else None


def list_subfolders(drive, parent_id):
    """Lista todas as subpastas de uma pasta."""
    q = f"mimeType='application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed=false"
    r = drive.files().list(q=q, fields="files(id,name)", orderBy="name").execute()
    return r.get("files", [])


def find_main_spreadsheet(drive, parent_id):
    """Encontra a planilha principal na pasta do cliente."""
    q = (
        f"(mimeType='application/vnd.google-apps.spreadsheet' or "
        f"mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') "
        f"and '{parent_id}' in parents and trashed=false"
    )
    r = drive.files().list(q=q, fields="files(id,name,mimeType)").execute()
    files = r.get("files", [])
    if not files:
        return None
    # Prefere arquivos que não sejam "PLANILHA DANI" (ajuste se precisar)
    return files[0]


def get_sheet_names(drive, sheets_service, file_info):
    """Retorna lista de nomes de abas de uma planilha."""
    fid  = file_info["id"]
    mime = file_info["mimeType"]

    if mime == "application/vnd.google-apps.spreadsheet":
        meta = sheets_service.spreadsheets().get(
            spreadsheetId=fid, fields="sheets.properties.title"
        ).execute()
        return [s["properties"]["title"] for s in meta.get("sheets", [])]
    else:
        request = drive.files().get_media(fileId=fid)
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = dl.next_chunk()
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names


def get_or_create_textos_folder(drive, client_folder_id):
    """Retorna (ou cria) a pasta TEXTOS dentro do cliente."""
    folder = find_folder_by_name(drive, "TEXTOS", client_folder_id)
    if folder:
        return folder["id"]
    meta = {
        "name": "TEXTOS",
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [client_folder_id],
    }
    new_folder = drive.files().create(body=meta, fields="id").execute()
    print("     📁 Pasta TEXTOS criada.")
    return new_folder["id"]


def list_existing_docs(drive, folder_id):
    """Retorna um set com os nomes dos arquivos já existentes em TEXTOS (com paginação)."""
    q = f"'{folder_id}' in parents and trashed=false"
    names = set()
    page_token = None
    while True:
        kwargs = dict(q=q, fields="nextPageToken, files(name)", pageSize=1000)
        if page_token:
            kwargs["pageToken"] = page_token
        r = drive.files().list(**kwargs).execute()
        for f in r.get("files", []):
            names.add(f["name"])
        page_token = r.get("nextPageToken")
        if not page_token:
            break
    return names


# ══════════════════════════════════════════════════════════
#  LEITURA DA PLANILHA
# ══════════════════════════════════════════════════════════

def detect_columns(header_rows):
    """Detecta índices de coluna a partir das linhas de cabeçalho da planilha.

    Varre as linhas de cabeçalho procurando palavras-chave conhecidas.
    Retorna um dict {campo: índice_coluna}. Usa _COL_DEFAULTS para campos
    não encontrados.
    """
    mapping = {}
    for row in header_rows:
        for col_idx, cell in enumerate(row):
            cell_norm = _norm(cell)
            if not cell_norm or cell_norm in ("none", "nan"):
                continue
            for field, keywords in _HEADER_KEYWORDS.items():
                if field not in mapping and any(kw in cell_norm for kw in keywords):
                    mapping[field] = col_idx
        if len(mapping) == len(_HEADER_KEYWORDS):
            break  # todos os campos encontrados

    # preenche campos não detectados com o fallback
    for field, default_idx in _COL_DEFAULTS.items():
        if field not in mapping:
            mapping[field] = default_idx

    return mapping


def _xlsx_cell(c):
    """Converte valor de célula xlsx para string, tratando datas formatadas como número."""
    if c is None:
        return ""
    if isinstance(c, (datetime.datetime, datetime.date)):
        from openpyxl.utils.datetime import to_excel
        dt = c if isinstance(c, datetime.datetime) else datetime.datetime(c.year, c.month, c.day)
        return str(int(to_excel(dt)))
    return str(c)


def _parse_rows(rows, col_map):
    """Converte linhas brutas em lista de produtos usando mapeamento dinâmico de colunas."""
    products = []
    for row in rows:
        if len(row) <= col_map["produto"]:
            continue

        def cell(idx):
            if idx is None or len(row) <= idx:
                return ""
            v = str(row[idx]).strip()
            return "" if v.lower() in ("none", "nan", "") else v

        produto   = cell(col_map["produto"])
        marca     = cell(col_map["marca"])
        aplicacao = cell(col_map["aplicacao"])
        codigo    = cell(col_map["codigo"]) or "SEM"

        if produto and produto.lower() not in ("produto", "none", "nan"):
            products.append({
                "produto":   produto,
                "marca":     marca,
                "aplicacao": aplicacao,
                "codigo":    codigo or "SEM",
            })
    return products


def read_products_from_spreadsheet(drive, sheets, file_info, sheets_filter=None):
    """Lê os produtos das abas da planilha (Google Sheets nativo ou .xlsx).

    sheets_filter: lista de nomes de abas a processar, ou None para todas.
    """
    fid   = file_info["id"]
    mime  = file_info["mimeType"]
    products = []

    if mime == "application/vnd.google-apps.spreadsheet":
        # ── Google Sheets nativo — busca todas as abas ────
        meta = sheets.spreadsheets().get(spreadsheetId=fid, fields="sheets.properties").execute()
        sheet_names = [s["properties"]["title"] for s in meta.get("sheets", [])]
        if sheets_filter:
            sheet_names = [s for s in sheet_names if s in sheets_filter]
        print(f"     📑 {len(sheet_names)} aba(s) selecionada(s): {', '.join(sheet_names)}")
        for sheet_name in sheet_names:
            # lê cabeçalhos (linhas 1 até LINHA_INICIO-1) para detectar colunas
            header_result = sheets.spreadsheets().values().get(
                spreadsheetId=fid,
                range=f"'{sheet_name}'!A1:Z{LINHA_INICIO - 1}"
            ).execute()
            col_map = detect_columns(header_result.get("values", []))
            print(f"       🗂  Colunas detectadas: produto={col_map['produto']} marca={col_map['marca']} "
                  f"codigo={col_map['codigo']} aplicacao={col_map['aplicacao']}")

            result = sheets.spreadsheets().values().get(
                spreadsheetId=fid,
                range=f"'{sheet_name}'!A{LINHA_INICIO}:Z2000"
            ).execute()
            rows = result.get("values", [])
            found = _parse_rows(rows, col_map)
            if found:
                print(f"       → Aba '{sheet_name}': {len(found)} produto(s)")
            products.extend(found)
    else:
        # ── Arquivo .xlsx — percorre todas as abas ────────
        request = drive.files().get_media(fileId=fid)
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = dl.next_chunk()
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
        all_names = list(wb.sheetnames)
        filtered  = [s for s in all_names if s in sheets_filter] if sheets_filter else all_names
        print(f"     📑 {len(filtered)} aba(s) selecionada(s): {', '.join(filtered)}")
        for sheet_name in filtered:
            ws = wb[sheet_name]
            # lê cabeçalhos para detectar colunas
            header_rows = []
            for row in ws.iter_rows(min_row=1, max_row=LINHA_INICIO - 1, values_only=True):
                header_rows.append([str(c) if c is not None else "" for c in row])
            col_map = detect_columns(header_rows)
            print(f"       🗂  Colunas detectadas: produto={col_map['produto']} marca={col_map['marca']} "
                  f"codigo={col_map['codigo']} aplicacao={col_map['aplicacao']}")

            rows = []
            for row in ws.iter_rows(min_row=LINHA_INICIO, values_only=True):
                rows.append([_xlsx_cell(c) for c in row])
            found = _parse_rows(rows, col_map)
            if found:
                print(f"       → Aba '{sheet_name}': {len(found)} produto(s)")
            products.extend(found)

    return products


# ══════════════════════════════════════════════════════════
#  GERAÇÃO VIA GEMINI
# ══════════════════════════════════════════════════════════

def _chamar_modelo(modelo, system_message, prompt):
    """Chama o modelo OpenAI."""
    if not OPENAI_API_KEY:
        raise Exception("OPENAI_API_KEY não configurada no .env")
    client = OpenAI(api_key=OPENAI_API_KEY)
    response = client.chat.completions.create(
        model=modelo,
        messages=[
            {"role": "system", "content": system_message},
            {"role": "user", "content": prompt},
        ],
        temperature=0.5,
    )
    return response.choices[0].message.content


def generate_with_groq(product, client_name):
    """Gera o texto via IA com rotação automática de modelos (Gemini + Groq)."""

    system_message = """Você é um especialista em SEO e copywriting para o Mercado Livre, criando anúncios de alta conversão para produtos industriais, automotivos e de abastecimento.

REGRAS ABSOLUTAS — NUNCA QUEBRE:
1. Responda SOMENTE com o conteúdo do anúncio. Zero introduções, zero "Aqui está", zero comentários.
2. NUNCA use asteriscos (*), hashtags (#), aspas ou qualquer markdown.
3. O título é uma frase SEO. NUNCA copie o nome bruto do produto como título.
4. Escreva o título APENAS UMA VEZ, na linha após o rótulo "TITULO".
5. NÃO existe seção VARIACAO. O anúncio tem apenas UM título.
6. Nos dois parágrafos de texto, NUNCA use "nosso", "nossa", "nós". Sempre terceira pessoa.
7. Na seção APLICACAO, use SOMENTE os dados fornecidos. NUNCA invente compatibilidades.
8. Os dois parágrafos de texto devem ser técnicos e específicos ao produto. NUNCA genéricos."""

    aplicacao = product['aplicacao'] if product['aplicacao'] else ''

    prompt = f"""Crie um anúncio completo para o Mercado Livre. Siga o modelo abaixo EXATAMENTE.

========================================================
EXEMPLO REAL DE ANUNCIO BEM FEITO (use como referência de qualidade):
========================================================

DADOS DE ENTRADA DO EXEMPLO:
Produto: FILTRO COMBUSTIVEL
Marca: AUTHOMIX
Codigo: FCO0507
Aplicacao: CHEVROLET:
S10 2.5 16V 14/

CITROEN: 
Aircross 1.6 16v 10/18, Berlingo 1.6
16v 05/07, Berlingo 1.8 8v 97/03, C3
1.2 12v 16/, C3 1.5 8v 14/16, C3 1.6
16v 03/14, C3 Picasso 1.6 16v 11/16,
C4 2.0 16v 09/13, C4 Lounge 1.6 16v
13/18, C4 Pallas 2.0 16v 07/13

FIAT:
Mobi 1.0 8v 16/, Uno 1.0 6v 16/17

HONDA:
Civic 1.8 16v 11/14, Civic 2.0 16v
13/17, Fit 1.4 16v 08/14, Fit 1.5 16v
14/

HYNDAI:
HB20 1.0 12V / 1.6 16V 12/19

KIA:
Sportage 2.0 16v 12/16

NISSAN:
Grand Livina 1.8 16v 09/14, Livina 1.6
16v 09/14, Kicks 1.6 16v 16/17, March
1.0 16v 11/16, Sentra 2.0 16v 06/16,
Versa 1.0 12v 15/, Tiida 1.8 16v 07/13,
Versa 1.6 16v 11/

PEGEOUT
106 1.0 8V 92/05, 08 1.6 16V 15/19,
206 1.4 8V 04/09, 205 1.4 8V 92/98,
206 1.6 16V 00/08, 206 1.6 8V 99/03,
207 1.4 8V 09/14, 207 1.6 16V 08/13,
208 1.5 8V 13/16, 208 1.6 16V 14/18,
3008 1.6 16V 10/17, 306 1.8 16V
97/04

RENAULT:
Clio 1.0 16v 01/13, Clio 1.0 8v 99/06,
Clio 1.6 16v 00/09, Clio 1.6 8v 98/02,
Duster 1.6 /2.0 16v 11/18, Fluence 1.6
16v 12/, Fluence 2.0 16v 11/, Gran
Tour 1.6 16v 06/12, Kangoo 1.6 16v
02/11

VOLKSWAGEN:
Cross Up! 1.0 12v 14/17, Gol 1.0 12v
16/18, Gol 1.6 8v 08/15, Golf 1.0 12v
16/18, Golf 1.6 16v 15/16, Golf 1.6
8v 06/14, Kombi 1.4 8v 06/12, Polo
1.6 8v 04/14, Polo 2.0 8v 08/12, Polo
Sedan 1.6 8v 04/14, Saveiro 1.6 16v
14/16

SAIDA CORRETA DO EXEMPLO:

TÍTULO (SEO Mercado Livre)

Filtro Combustível Gol Voyage Polo Civic Hb20 S10

DESCRIÇÃO COMPLETA (Padrão Escalada Ecom)
--------------------------------------------------
O QUE VEM NA CAIXA
- 01 Filtro de Combustível
- Marca: AUTHOMIX
- Código/Referência: FCO0507
--------------------------------------------------
APLICAÇÃO

CHEVROLET:
- S10 2.5 16V 2014/

CITROËN:
- Aircross 1.6 16V 2010/2018
- Berlingo 1.6 16V 2005/2007
- Berlingo 1.8 8V 1997/2003
- C3 1.2 12V 2016/
- C3 1.5 8V 2014/2016
- C3 1.6 16V 2003/2014
- C3 Picasso 1.6 16V 2011/2016
- C4 2.0 16V 2009/2013
- C4 Lounge 1.6 16V 2013/2018
- C4 Pallas 2.0 16V 2007/2013

FIAT:
- Mobi 1.0 8V 2016/
- Uno 1.0 2016/2017

HONDA:
- Civic 1.8 16V 2011/2014
- Civic 2.0 16V 2013/2017
- Fit 1.4 16V 2008/2014
- Fit 1.5 16V 2014/

HYUNDAI:
- HB20 1.0 12V / 1.6 16V 2012/2019

KIA:
- Sportage 2.0 16V 2012/2016

NISSAN:
- Grand Livina 1.8 16V 2009/2014
- Livina 1.6 16V 2009/2014
- Kicks 1.6 16V 2016/2017
- March 1.0 16V 2011/2016
- Sentra 2.0 16V 2006/2016
- Versa 1.0 12V 2015/
- Tiida 1.8 16V 2007/2013
- Versa 1.6 16V 2011/

PEUGEOT:
- 106 1.0 8V 1992/2005
- 208 1.6 16V 2014/2018
- 3008 1.6 16V 2010/2017
- 206, 207, 306 e demais modelos conforme aplicação

RENAULT:
- Clio 1.0 16V 2001/2013
- Clio 1.0 8V 1999/2006
- Duster 1.6 / 2.0 16V 2011/2018
- Fluence 1.6 / 2.0 16V
- Kangoo 1.6 16V 2002/2011

VOLKSWAGEN:
- Cross Up! 1.0 12V 2014/2017
- Gol 1.0 12V 2016/2018
- Gol 1.6 8V 2008/2015
- Golf 1.0 / 1.6 2006/2018
- Kombi 1.4 8V 2006/2012
- Polo 1.6 / 2.0 2004/2014
- Polo Sedan 1.6 8V 2004/2014
- Saveiro 1.6 16V 2014/2016
--------------------------------------------------
Mantenha o sistema de alimentação do motor protegido com o filtro de combustível AUTHOMIX. Desenvolvido para reter impurezas presentes no combustível, ajuda a preservar bicos injetores, bomba de combustível e demais componentes do sistema, contribuindo para melhor desempenho e funcionamento do veículo.

A substituição periódica do filtro auxilia na prevenção de falhas, melhora a eficiência do motor e ajuda a manter o consumo adequado de combustível. Produzido com materiais de qualidade, oferece excelente filtragem e maior confiabilidade para o dia a dia.
--------------------------------------------------
INSTITUCIONAL

A EUnaPEÇAS trabalha com peças e acessórios automotivos de qualidade, oferecendo produtos confiáveis para manutenção, segurança e desempenho do seu veículo.

Nosso compromisso é entregar produtos de procedência, envio rápido e atendimento de confiança, ajudando você a manter seu veículo sempre em excelente funcionamento.


========================================================
REGRAS DO TITULO (aprenda com o exemplo):
- O título não é o nome do produto. É uma mistura coerencte de palavras-chave de busca otimizada.
- Inclua especificações técnicas e/ou modelos compatíveis extraídos da APLICACAO.
- Máximo 60 caracteres. Sem a marca. Sem o código.
- RUIM: "LAMPADA H7 24V 100W CAMINHAO"  (nome bruto)
- BOM: "Lampada H7 24V 100W Caminhao Truck Par Alta Potencia"  (especificações + contexto de busca)
- RUIM: "FILTRO COMBUSTIVEL"  (genérico demais)
- BOM: "Filtro Combustivel Gol Voyage Polo Clio Fit Civic 1.0 1.4"  (modelos compatíveis)

========================================================
REGRAS DA SECAO APLICACAO (aprenda com o exemplo):
- Para produtos com aplicação vazia: liste usos, compatibilidades e especificações técnicas fazendo buscas do produto na internet.
- Para autopeças com lista de veículos: agrupe por montadora e busque ano, modelo e  motor para cada um. Exemplo:
  Volkswagen: 
  - Gol 2010-2020 1.0 1.4 2.0
  - Voyage 2010-2020 1.0 1.4 2.0
  - Polo Golf 2010-2020 1.0 1.4 2.0
  Fiat: 
  - Uno 2010-2020 1.0 1.4
  - Palio 2010-2020 1.0 1.4
  - Siena 2010-2020 1.0 1.4
  Honda: 
  - Civic 2010-2020 1.0 1.4
  - Fit 2010-2020 1.0 1.4

- NUNCA escreva o código do produto na APLICACAO.

========================================================
REGRAS DOS PARAGRAFOS DE TEXTO (aprenda com o exemplo):
- Parágrafo 1: explique o que é o produto, para que serve e qual problema resolve. 3 a 4 frases.
- Parágrafo 2: detalhe materiais, especificações técnicas, certificações e diferenciais. 3 a 4 frases.
- Ambos em terceira pessoa. Sem bullet points. Sem "nosso/nossa".
- NUNCA escreva parágrafos genéricos que sirvam para qualquer produto.

========================================================
AGORA CRIE O ANUNCIO PARA:

DADOS DO PRODUTO:
Produto: {product['produto']}
Marca: {product['marca']}
Codigo: {product['codigo']}
Aplicacao: {aplicacao if aplicacao else 'Não informada'}
Loja: {client_name}

FORMATO OBRIGATORIO — copie os rótulos EXATAMENTE como no exemplo:

TÍTULO (SEO Mercado Livre)

[uma unica linha de titulo SEO]

DESCRICAO COMPLETA (Padrao Ecom)
--------------------------------------------------
O QUE VEM NA CAIXA
- {product['produto']}
- Marca: {product['marca']}
- Codigo/Referencia: {product['codigo']}
--------------------------------------------------
APLICACAO
[bullets de aplicação/compatibilidade baseados nos dados fornecidos]
--------------------------------------------------
[Paragrafo 1 — específico ao produto, explica uso e benefício]

[Paragrafo 2 — materiais, especificações técnicas, diferenciais]
--------------------------------------------------
INSTITUCIONAL

A {client_name} oferece produtos de qualidade para abastecimento, manutenção e operação, disponibilizando soluções confiáveis para maior segurança e eficiência no dia a dia.

Nosso compromisso é entregar produtos de procedência, envio rápido e atendimento de confiança, ajudando sua operação a manter desempenho, organização e segurança."""

    MAX_TENTATIVAS = 4
    for tentativa in range(1, MAX_TENTATIVAS + 1):
        try:
            return _chamar_modelo(OPENAI_MODEL, system_message, prompt)
        except (APIConnectionError, APITimeoutError) as e:
            espera = 15
            print(f"\n  ⏳ Erro de conexão ({str(e)[:50]}...) — aguardando {espera}s "
                  f"(tentativa {tentativa}/{MAX_TENTATIVAS})...")
            time.sleep(espera)
        except Exception as e:
            erro = str(e)
            if "429" in erro or "rate_limit" in erro.lower():
                match = re.search(r'try again in ([\d.]+)s', erro, re.IGNORECASE)
                espera = int(float(match.group(1))) + 5 if match else 60
                print(f"\n  ⏳ Rate limit — aguardando {espera}s "
                      f"(tentativa {tentativa}/{MAX_TENTATIVAS})...")
                time.sleep(espera)
            elif any(k in erro.lower() for k in ("connect", "timeout", "read", "conexão", "interrompida", "connection")):
                espera = 15
                print(f"\n  ⏳ Erro de conexão ({erro[:50]}...) — aguardando {espera}s "
                      f"(tentativa {tentativa}/{MAX_TENTATIVAS})...")
                time.sleep(espera)
            else:
                print(f"\n  ❌ Erro inesperado: {erro}")
                raise
    raise Exception(f"Falhou após {MAX_TENTATIVAS} tentativas.")


# ══════════════════════════════════════════════════════════
#  CRIAÇÃO DO GOOGLE DOC
# ══════════════════════════════════════════════════════════

def create_google_doc(drive, docs, title, content, folder_id):
    """Cria um Google Doc com o conteúdo gerado e salva na pasta TEXTOS."""
    # Cria o documento vazio na pasta certa
    meta = {
        "name": title,
        "mimeType": "application/vnd.google-apps.document",
        "parents": [folder_id],
    }
    doc = drive.files().create(body=meta, fields="id").execute()
    doc_id = doc["id"]

    # Garante que haja texto para inserir, evitando erro 400 da API do Docs
    if not content or not str(content).strip():
        content = "[ERRO: A inteligência artificial não retornou nenhum texto para este anúncio. Verifique o prompt ou os dados do produto.]"

    # Insere o texto
    docs.documents().batchUpdate(
        documentId=doc_id,
        body={"requests": [{"insertText": {"location": {"index": 1}, "text": content}}]},
    ).execute()

    return doc_id


# ══════════════════════════════════════════════════════════
#  PROCESSAMENTO POR CLIENTE
# ══════════════════════════════════════════════════════════

def process_client(client_name, client_folder_id, drive, sheets, docs, sheets_filter=None):
    """Processa todos os produtos de um cliente."""
    sep = "─" * 55
    print(f"\n{sep}")
    print(f"  Cliente: {client_name}")
    print(sep)

    # Planilha
    spreadsheet = find_main_spreadsheet(drive, client_folder_id)
    if not spreadsheet:
        print("  ⚠️  Nenhuma planilha encontrada. Pulando.")
        return 0, 0

    print(f"  📊 Planilha: {spreadsheet['name']}")

    # Lê produtos
    try:
        products = read_products_from_spreadsheet(drive, sheets, spreadsheet, sheets_filter=sheets_filter)
    except Exception as e:
        print(f"  ❌ Erro ao ler planilha: {e}")
        return 0, 0

    print(f"  📦 {len(products)} produtos encontrados")
    if not products:
        return 0, 0

    # Mostra os primeiros nomes lidos da planilha para conferência
    for p in products[:5]:
        print(f"       → {p['produto'][:60]}")
    if len(products) > 5:
        print(f"       ... e mais {len(products) - 5}")

    # Pasta TEXTOS
    textos_id = get_or_create_textos_folder(drive, client_folder_id)
    existing  = list_existing_docs(drive, textos_id)
    print(f"  📝 {len(existing)} docs já existentes em TEXTOS (serão pulados)")

    # Se houver docs existentes, mostra os primeiros para diagnóstico
    if existing:
        sample = sorted(existing)[:5]
        for name in sample:
            print(f"       já existe: {name[:60]}")
        if len(existing) > 5:
            print(f"       ... e mais {len(existing) - 5}")

    created = skipped = errors = 0

    # Rastreia quantas vezes cada nome de produto aparece para gerar títulos únicos
    name_seen: dict = defaultdict(int)

    for i, product in enumerate(products, 1):
        base_title = f"{product['produto']} - {product['codigo']}"
        name_seen[base_title] += 1
        occurrence = name_seen[base_title]
        title = base_title if occurrence == 1 else f"{base_title} ({occurrence})"

        prefix = f"  [{i:>3}/{len(products)}]"

        if title in existing:
            print(f"{prefix} ⏭️  Já existe — {title[:45]}")
            skipped += 1
            continue

        print(f"{prefix} ✨ Gerando  — {title[:45]}...", end="", flush=True)
        try:
            content = generate_with_groq(product, client_name)
            create_google_doc(drive, docs, title, content, textos_id)
            print(" ✅")
            created += 1
            existing.add(title)           # marca como criado para esta sessão
            time.sleep(DELAY_ENTRE_PRODUTOS)
        except Exception as e:
            print(f" ❌ {e}")
            errors += 1

    print(f"\n  Resultado: {created} criados | {skipped} pulados | {errors} erros")
    return created, skipped


# ══════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════

def main():
    print("\n" + "═" * 55)
    print("  AUTOMAÇÃO MERCADO LIVRE — Geração de Anúncios")
    print("═" * 55)

    print(f"\n🤖 Modelo de IA: {OPENAI_MODEL}  [OpenAI]")

    # Autenticação
    print("\n🔑 Autenticando com Google...")
    creds  = get_credentials()
    drive  = build("drive",  "v3", credentials=creds)
    sheets = build("sheets", "v4", credentials=creds)
    docs   = build("docs",   "v1", credentials=creds)
    print("✅ Autenticado!\n")

    # Pasta raiz — ID fixo
    print(f"📁 Usando pasta MERCADO LIVRE (ID: {PASTA_RAIZ_ID})")

    # Lista clientes
    clients = list_subfolders(drive, PASTA_RAIZ_ID)
    if not clients:
        print("⚠️  Nenhuma subpasta de cliente encontrada.")
        return

    print(f"\n👥 {len(clients)} clientes disponíveis:")
    for i, c in enumerate(clients, 1):
        print(f"   {i:>2}. {c['name']}")

    # Seleção
    print("\nQual cliente processar?")
    print("   0 = Todos os clientes")
    choice = input("Digite o número: ").strip()

    if choice == "0":
        selected = clients
    else:
        try:
            idx = int(choice) - 1
            selected = [clients[idx]]
        except (ValueError, IndexError):
            print("❌ Opção inválida.")
            return

    # Processa
    total_criados = total_pulados = 0
    for client in selected:
        c, s = process_client(client["name"], client["id"], drive, sheets, docs)
        total_criados  += c
        total_pulados  += s

    # Resumo final
    print("\n" + "═" * 55)
    print("  CONCLUÍDO!")
    print(f"  Documentos criados : {total_criados}")
    print(f"  Produtos pulados   : {total_pulados}")
    print("═" * 55 + "\n")


if __name__ == "__main__":
    main()